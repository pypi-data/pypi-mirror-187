"""
OwnCloud compatible cloud services (e.g. NextCloud) often offer good support for collaborative
editing of tabular data in an online replica of Excel.

Public read-only shares provide a simple way to access such data programmatically.
"""
import typing
import pathlib
import tempfile
import collections

from csvw.dsv import UnicodeWriter
from collabutils.util import warn
try:
    import owncloud
    import openpyxl
except ImportError as e:  # pragma: no cover
    warn('oc.Spreadsheet', 'owncloud', e)

from collabutils.base import SharedSpreadsheetMixin

__all__ = ['Spreadsheet']

File = collections.namedtuple('File', 'name path content_type'.split())


class Share:
    def __init__(self, link):
        self.client = owncloud.Client.from_public_link(link)

    @property
    def files(self):
        return [File(f.name, f.path, f.get_content_type()) for f in self.client.list('')]


class Spreadsheet(SharedSpreadsheetMixin):
    """
    Typical usage in a `cldfbench.Dataset`'s `cmd_download` method:

    .. code-block:: python

        >>> document = Spreadsheet(
        ...     'test.xlsx', 'https://share.eva.mpg.de/index.php/s/pGaomxWqHPqxeEA')
        >>> document.fetch_sheets(
        ...     sheets={'Varieties': 'languages.tsv'})
        ...     outdir=dataset.etc_dir,
        ...     delimiter='\t')
    """
    def __init__(self, fname, share_link):
        share = Share(share_link)
        for f in share.files:
            if f.name == fname:
                self.file = f
                break
        else:
            raise ValueError(fname)
        self.client = share.client

    def fetch_sheets(
            self,
            sheets: typing.Optional[typing.Dict[str, str]] = None,
            outdir: typing.Optional[typing.Union[pathlib.Path, str]] = '.',
            **kw,
    ):
        import shutil

        def _excel_value(x):
            if x is None:
                return ""
            if isinstance(x, float) and int(x) == x:
                # Since Excel does not have an integer type, integers are rendered as "n.0",
                # which in turn confuses type detection of tools like csvkit. Thus, we normalize
                # numbers of the form "n.0" to "n".
                return '{0}'.format(int(x))  # pragma: no cover
            return '{0}'.format(x).strip()

        with tempfile.TemporaryDirectory() as tmp:
            tmppath = str(pathlib.Path(tmp) / self.file.name)
            self.client.get_file(self.file.path, tmppath)
            shutil.copy(tmppath, 'lexibank.xlsx')
            outdir = outdir or self
            wb = openpyxl.load_workbook(tmppath, data_only=True)
            print(wb.sheetnames)
            for i, sname in enumerate(wb.sheetnames, start=1):
                if sheets is None or (sname in sheets):
                    fname = 'sheet_{}.csv'.format(i) if sheets is None else sheets[sname]
                    sheet = wb[sname]
                    with UnicodeWriter(pathlib.Path(outdir) / fname, **kw) as writer:
                        for row in sheet.rows:
                            writer.writerow([_excel_value(col.value) for col in row])
