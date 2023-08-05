from typing import Optional, Dict, List

import openpyxl
from openpyxl.utils import rows_from_range, coordinate_to_tuple
from slugify import slugify


def clean_string(value: str):
    if isinstance(value, str):
        return value.strip()


def calcule_max_row(ws, row=1, col=1):
    while True:
        value = ws.cell(row=row, column=col).value
        value_cleaned = clean_string(value)
        if not value_cleaned:
            break
        row += 1
    return row - 1


def calcule_max_col(ws, row=1, col=1):
    while True:
        value = ws.cell(row=row, column=col).value
        value_cleaned = clean_string(value)
        if not value_cleaned:
            break
        col += 1
    return col - 1


def get_first_cell(ws, interval='A1:J10'):
    cells = []
    for rows in rows_from_range(interval):
        for cell_address in rows:
            cells.append(cell_address)
    for cell_address in cells:
        value = ws[cell_address].value
        value_cleaned = clean_string(value)
        if value_cleaned:
            return coordinate_to_tuple(cell_address)


def get_last_cell(ws, min_row, min_col):
    max_row = calcule_max_row(ws, row=min_row, col=min_col)
    max_col = calcule_max_col(ws, row=min_row, col=min_col)
    return max_row, max_col


def xlsx_to_dict(filename: str, sheetname: Optional[str] = None, read_only=True) -> List[Dict]:
    fields = []
    data = []

    wb = openpyxl.load_workbook(filename, read_only=read_only)
    ws = wb[sheetname or wb.sheetnames[0]]

    min_row, min_col = get_first_cell(ws)
    max_row, max_col = get_last_cell(ws, min_row, min_col)

    for x, itens in enumerate(ws.iter_rows(values_only=True, min_row=min_row, min_col=min_col, max_row=max_row,
                                           max_col=max_col)):
        if x == 0:
            # fields
            fields.extend([slugify(item).replace('-', '_') for item in itens])
        else:
            # values
            data.append({k: v for k, v in zip(fields, itens)})

    wb.close()

    return data
