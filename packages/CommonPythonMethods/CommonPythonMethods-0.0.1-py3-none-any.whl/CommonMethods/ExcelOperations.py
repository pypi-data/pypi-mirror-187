import openpyxl


def Fetch_Cell_Values_From_ExcelSheets(FilePath, SheetName, Row, Column):
    """Fetch Excel Cell values for respective sheet """
    wk = openpyxl.load_workbook(FilePath)
    wk.active = wk[SheetName]
    sheet = wk.active
    row = int(Row)
    col = int(Column)
    data = sheet.cell(row, col)
    data1 = data.value
    return str(data1)


def Fetch_Excel_Coloumn_Header(FilePath, SheetName):  # pylint:disable=invalid-name
    """ExcelOperations"""
    wk = openpyxl.load_workbook(FilePath)
    wk.active = wk[SheetName]
    sheet = wk.active
    headerValues = []
    maxCol = sheet.max_column
    for i in range(1, maxCol + 1):
        data = sheet.cell(1, i)
        headerValues.append(data.value)
        print(data.value)
    return headerValues


def Excel_StrToInt_Conversion(FilePath, cell, data):
    """ExcelOperations"""
    wk = openpyxl.load_workbook(FilePath)
    sheet = wk.active
    sheet[cell] = int(data)
    wk.save(FilePath)


def Excel_IntToStr_Conversion(FilePath, cell, data):
    """ExcelOperations"""
    wk = openpyxl.load_workbook(FilePath)
    sheet = wk.active
    sheet[cell] = int(data)
    wk.save(FilePath)
